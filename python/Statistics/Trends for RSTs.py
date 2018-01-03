from openpyxl import Workbook, load_workbook
from time import strptime

seasonal_trends = True  # False for monthly trends

output_dir = 'C:/Users/hatzv/Documents/Geography/RSTs/python/Statistics/Results/'

# Prepare the worksheets for the NCEP output file
if seasonal_trends:
    output_excel_filename_NCEP = output_dir + 'Seasonal_trends_RSTs_NCEP_1979-2016.xlsx'
else:
    output_excel_filename_NCEP = output_dir + 'Monthly_trends_RSTs_NCEP_1979-2016.xlsx'
wb_NCEP_trends = Workbook()
ws_NCEP_NO_RSTs = wb_NCEP_trends.get_sheet_by_name("Sheet")
ws_NCEP_NO_RSTs.title = "NO_RSTs"
ws_NCEP_East = wb_NCEP_trends.create_sheet("East")
ws_NCEP_West = wb_NCEP_trends.create_sheet("West")
ws_NCEP_Central = wb_NCEP_trends.create_sheet("Central")
ws_NCEP_All = wb_NCEP_trends.create_sheet("All RSTs")

# Prepare the worksheets for the ERA output file
if seasonal_trends:
    output_excel_filename_ERA = output_dir + 'Seasonal_trends_RSTs_ERA_1979-2016.xlsx'
else:
    output_excel_filename_ERA = output_dir + 'Monthly_trends_RSTs_ERA_1979-2016.xlsx'
wb_ERA_trends = Workbook()
ws_ERA_NO_RSTs = wb_ERA_trends.get_sheet_by_name("Sheet")
ws_ERA_NO_RSTs.title = "NO_RSTs"
ws_ERA_East = wb_ERA_trends.create_sheet("East")
ws_ERA_West = wb_ERA_trends.create_sheet("West")
ws_ERA_Central = wb_ERA_trends.create_sheet("Central")
ws_ERA_All = wb_ERA_trends.create_sheet("All RSTs")

# Prepare the worksheets for the ERA_2_5 output file
if seasonal_trends:
    output_excel_filename_ERA_2_5 = output_dir + 'Seasonal_trends_RSTs_ERA_2_5_1979-2016.xlsx'
else:
    output_excel_filename_ERA_2_5 = output_dir + 'Monthly_trends_RSTs_ERA_2_5_1979-2016.xlsx'
wb_ERA_2_5_trends = Workbook()
ws_ERA_2_5_NO_RSTs = wb_ERA_2_5_trends.get_sheet_by_name("Sheet")
ws_ERA_2_5_NO_RSTs.title = "NO_RSTs"
ws_ERA_2_5_East = wb_ERA_2_5_trends.create_sheet("East")
ws_ERA_2_5_West = wb_ERA_2_5_trends.create_sheet("West")
ws_ERA_2_5_Central = wb_ERA_2_5_trends.create_sheet("Central")
ws_ERA_2_5_All = wb_ERA_2_5_trends.create_sheet("All RSTs")

# Prepare the output tables for all created sheets
table_NCEP_NO_RSTs = ws_NCEP_NO_RSTs['B2':'AM13']
table_NCEP_East = ws_NCEP_East['B2':'AM13']
table_NCEP_West = ws_NCEP_West['B2':'AM13']
table_NCEP_Central = ws_NCEP_Central['B2':'AM13']
table_NCEP_All = ws_NCEP_All['B2':'AM5']

table_ERA_NO_RSTs = ws_ERA_NO_RSTs['B2':'AM13']
table_ERA_East = ws_ERA_East['B2':'AM13']
table_ERA_West = ws_ERA_West['B2':'AM13']
table_ERA_Central = ws_ERA_Central['B2':'AM13']
table_ERA_All = ws_ERA_All['B2':'AM5']

table_ERA_2_5_NO_RSTs = ws_ERA_2_5_NO_RSTs['B2':'AM13']
table_ERA_2_5_East = ws_ERA_2_5_East['B2':'AM13']
table_ERA_2_5_West = ws_ERA_2_5_West['B2':'AM13']
table_ERA_2_5_Central = ws_ERA_2_5_Central['B2':'AM13']
table_ERA_2_5_All = ws_ERA_2_5_All['B2':'AM5']

# Prepare the worksheets for the input classifications
excel_filename_NCEP = 'C:/Users/hatzv/Documents/Geography/RSTs/python/Statistics/Results/RST_classification_NCEP_1979-2016.xlsx'
excel_filename_ERA = 'C:/Users/hatzv/Documents/Geography/RSTs/python/Statistics/Results/RST_classification_ERA_1979-2016.xlsx'
excel_filename_ERA_2_5 = 'C:/Users/hatzv/Documents/Geography/RSTs/python/Statistics/Results/RST_classification_ERA_2.5_1979-2016.xlsx'

wb_NCEP = load_workbook(excel_filename_NCEP, read_only=True)
ws_NCEP = wb_NCEP.active
table_NCEP = ws_NCEP['A2':'AM367']

wb_ERA = load_workbook(excel_filename_ERA, read_only=True)
ws_ERA = wb_ERA.active
table_ERA = ws_ERA['A2':'AM367']

wb_ERA_2_5 = load_workbook(excel_filename_ERA_2_5, read_only=True)
ws_ERA_2_5 = wb_ERA_2_5.active
table_ERA_2_5 = ws_ERA_2_5['A2':'AM367']

for row in range(1, ws_NCEP.max_row-1):
    for col in range(1, ws_NCEP.max_column):
        any_NCEP_RST = True  # A flag for marking if an RST found or not
        any_ERA_RST = True  # A flag for marking if an RST found or not
        any_ERA_2_5_RST = True  # A flag for marking if an RST found or not

        # Find the output row
        current_month = strptime(table_NCEP[row][0].value, '%b').tm_mon
        if seasonal_trends:
            if current_month == 12 or current_month <=2:  # DJF
                current_row = 0
            elif current_month >= 3 and current_month <=5:  # MAM
                current_row = 1
            elif current_month >= 6 and current_month <= 8:  # JJA
                current_row = 2
            else:  # SON
                current_row = 3
        else:
            current_row = current_month - 1

        NCEP_value = table_NCEP[row][col].value
        ERA_value = table_ERA[row][col].value
        ERA_2_5_value = table_ERA_2_5[row][col].value

        if NCEP_value == "No RST":
            output_NCEP_table = table_NCEP_NO_RSTs
            any_NCEP_RST = False
        elif NCEP_value == "East":
            output_NCEP_table = table_NCEP_East
        elif NCEP_value == "Central":
            output_NCEP_table = table_NCEP_Central
        elif NCEP_value == "West":
            output_NCEP_table = table_NCEP_West

        if ERA_value == "No RST":
            output_ERA_table = table_ERA_NO_RSTs
            any_ERA_RST = False
        elif ERA_value == "East":
            output_ERA_table = table_ERA_East
        elif ERA_value == "Central":
            output_ERA_table = table_ERA_Central
        elif ERA_value == "West":
            output_ERA_table = table_ERA_West

        if ERA_2_5_value == "No RST":
            output_ERA_2_5_table = table_ERA_2_5_NO_RSTs
            any_ERA_2_5_RST = False
        elif ERA_2_5_value == "East":
            output_ERA_2_5_table = table_ERA_2_5_East
        elif ERA_2_5_value == "Central":
            output_ERA_2_5_table = table_ERA_2_5_Central
        elif ERA_2_5_value == "West":
            output_ERA_2_5_table = table_ERA_2_5_West

        # TODO: Refactor this entire ugly thing.
        if NCEP_value is not None:  # Make sure not to read the None value in a non-leap year
            current_output_value_NCEP = output_NCEP_table[current_row][col-1].value
            if current_output_value_NCEP is not None:
                output_NCEP_table[current_row][col - 1].value = current_output_value_NCEP + 1
            else:
                output_NCEP_table[current_row][col - 1].value = 1

            if any_NCEP_RST:
                output_NCEP_table = table_NCEP_All
                current_output_value_NCEP = output_NCEP_table[current_row][col-1].value
                if current_output_value_NCEP is not None:
                    output_NCEP_table[current_row][col - 1].value = current_output_value_NCEP + 1
                else:
                    output_NCEP_table[current_row][col - 1].value = 1

            if ERA_value is not None:  # Make sure not to read the None value in a non-leap year
                current_output_value_ERA = output_ERA_table[current_row][col - 1].value
                if current_output_value_ERA is not None:
                    output_ERA_table[current_row][col - 1].value = current_output_value_ERA + 1
                else:
                    output_ERA_table[current_row][col - 1].value = 1

            if any_ERA_RST:
                output_ERA_table = table_ERA_All
                current_output_value_ERA = output_ERA_table[current_row][col-1].value
                if current_output_value_ERA is not None:
                    output_ERA_table[current_row][col - 1].value = current_output_value_ERA + 1
                else:
                    output_ERA_table[current_row][col - 1].value = 1

            if ERA_2_5_value is not None:  # Make sure not to read the None value in a non-leap year
                current_output_value_ERA_2_5 = output_ERA_2_5_table[current_row][col - 1].value
                if current_output_value_ERA_2_5 is not None:
                    output_ERA_2_5_table[current_row][col - 1].value = current_output_value_ERA_2_5 + 1
                else:
                    output_ERA_2_5_table[current_row][col - 1].value = 1

            if any_ERA_2_5_RST:
                output_ERA_2_5_table = table_ERA_2_5_All
                current_output_value_ERA_2_5 = output_ERA_2_5_table[current_row][col-1].value
                if current_output_value_ERA_2_5 is not None:
                    output_ERA_2_5_table[current_row][col - 1].value = current_output_value_ERA_2_5 + 1
                else:
                    output_ERA_2_5_table[current_row][col - 1].value = 1

wb_NCEP_trends.save(output_excel_filename_NCEP)
wb_NCEP_trends.save(output_excel_filename_ERA)
wb_NCEP_trends.save(output_excel_filename_ERA_2_5)
