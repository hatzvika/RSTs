from openpyxl import load_workbook, Workbook

# Prepare the worksheet for the output comparison
excel_filename_comparison = 'C:/Users/hatzv/Documents/Geography/RSTs/python/Statistics/Results/classification_comparison_1979-2016.xlsx'
wb_comparison = Workbook()
ws_comparison = wb_comparison.active
table_comparison = ws_comparison['B2':'T15']

# Prepare the worksheets for the input classifications
excel_filename_NCEP = 'C:/Users/hatzv/Documents/Geography/RSTs/python/Statistics/Results/RST_classification_NCEP_1979-2016.xlsx'
excel_filename_ERA = 'C:/Users/hatzv/Documents/Geography/RSTs/python/Statistics/Results/RST_classification_ERA_1979-2016.xlsx'
excel_filename_ERA_2_5 = 'C:/Users/hatzv/Documents/Geography/RSTs/python/Statistics/Results/RST_classification_ERA_2.5_1979-2016.xlsx'
excel_filename_Isabella = 'C:/Users/hatzv/Documents/Geography/Research_help/synoptic_classification/My_classification/synoptic_classification_1948-21_Sep_2017.xlsx'

wb_NCEP = load_workbook(excel_filename_NCEP, read_only=True)
ws_NCEP = wb_NCEP.active
table_NCEP = ws_NCEP['B2':'AM367']

wb_ERA = load_workbook(excel_filename_ERA, read_only=True)
ws_ERA = wb_ERA.active
table_ERA = ws_ERA['B2':'AM367']

wb_ERA_2_5 = load_workbook(excel_filename_ERA_2_5, read_only=True)
ws_ERA_2_5 = wb_ERA_2_5.active
table_ERA_2_5 = ws_ERA_2_5['B2':'AM367']

wb_Isabella = load_workbook(excel_filename_Isabella, read_only=True)
ws_Isabella = wb_Isabella.active
table_Isabella = ws_Isabella['AH2':'BS367']

for row in range (ws_NCEP.max_row - 1):
    for col in range (ws_NCEP.max_column - 1):
        NCEP_value = table_NCEP[row][col].value
        ERA_value = table_ERA[row][col].value
        ERA_2_5_value = table_ERA_2_5[row][col].value
        isabella_value = table_Isabella[row][col].value

        if NCEP_value == "No RST":
            output_row_NCEP = 0
        elif NCEP_value == "East":
            output_row_NCEP = 1
        elif NCEP_value == "West":
            output_row_NCEP = 2
        elif NCEP_value == "Central":
            output_row_NCEP = 3

        if ERA_value == "No RST":
            output_row_ERA = 5
        elif ERA_value == "East":
            output_row_ERA = 6
        elif ERA_value == "West":
            output_row_ERA = 7
        elif ERA_value == "Central":
            output_row_ERA = 8

        if ERA_2_5_value == "No RST":
            output_row_ERA_2_5 = 10
        elif ERA_2_5_value == "East":
            output_row_ERA_2_5 = 11
        elif ERA_2_5_value == "West":
            output_row_ERA_2_5 = 12
        elif ERA_2_5_value == "Central":
            output_row_ERA_2_5 = 13

        if isabella_value is not None:  # Make sure not to read the None value in a non-leap year
            current_output_value_NCEP = table_comparison[output_row_NCEP][isabella_value-1].value
            if current_output_value_NCEP is not None:
                table_comparison[output_row_NCEP][isabella_value-1].value = current_output_value_NCEP +1
            else:
                table_comparison[output_row_NCEP][isabella_value-1].value = 1

            current_output_value_ERA = table_comparison[output_row_ERA][isabella_value - 1].value
            if current_output_value_ERA is not None:
                table_comparison[output_row_ERA][isabella_value - 1].value = current_output_value_ERA + 1
            else:
                table_comparison[output_row_ERA][isabella_value - 1].value = 1

            current_output_value_ERA_2_5 = table_comparison[output_row_ERA_2_5][isabella_value - 1].value
            if current_output_value_ERA_2_5 is not None:
                table_comparison[output_row_ERA_2_5][isabella_value - 1].value = current_output_value_ERA_2_5 + 1
            else:
                table_comparison[output_row_ERA_2_5][isabella_value - 1].value = 1

wb_comparison.save(excel_filename_comparison)



