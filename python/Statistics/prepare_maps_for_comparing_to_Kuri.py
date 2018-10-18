import matplotlib.pyplot as plt
from python.Plot_RSTs.Plot_RSTs import PlotRSTs
from openpyxl import load_workbook

show_info = False
save_maps = True

# Parameters for the calculations that can change
use_interpolation = True
only_longest_separate = True
polyfit_rst = False

# Parameters for the calculations that can't change
data_to_map_var = 'Geostrophic Vorticity'
show_dots = False

start_year = 2000
end_year = 2005

year_list = [str(x) for x in range(start_year, end_year+1)]

# make a list of all desired dates (those for comparison)
wb = load_workbook('C:/Users/hatzv/Documents/Geography/RSTs/python/Statistics/Results/Kuri_synoptic_classification_2000_2004.xlsx', data_only=True) # data_only means getting the real values and not the functions in a cell
ws = wb['Sheet1']
dates_excel = ws['A3:A1829']
# partial_NCEP = ws['F3:F1829']  # Kuri
# partial_Kuri = ws['H3:H1829']  # Kuri
partial_NCEP = ws['M3:M1829']  # Pinhas
partial_Kuri = ws['O3:O1829']  # Pinhas

desired_dates_list= []
for excel_day in range(0,len(dates_excel)):
    if (partial_NCEP[excel_day][0].value == 0) or (partial_Kuri[excel_day][0].value == 0):
        desired_dates_list.append(dates_excel[excel_day][0].value)

maps_counter = 0
for current_year in year_list:
    print(current_year)
    plotRSTs_instance = PlotRSTs('NCEP', current_year)
    data_string_time = plotRSTs_instance.data_string_time
    previous_month_day = ""
    leap_year_offset = 0
    rows_counter = 2
    #for current_day in data_string_time[180:-1]:
    for current_day in data_string_time:
        current_day_str = str(current_day)
        if str(desired_dates_list[maps_counter])[0:10] == current_day_str[0:10]:
            maps_counter = maps_counter + 1
            print(current_day_str)

            daily_rst_classification, slp_data, geostrophic_vorticity_map, is_rst_condition_met = plotRSTs_instance.calculate_maps_data(
                current_day_str,
                use_interpolation=use_interpolation,
                data_to_map=data_to_map_var,
                show_dots=show_dots,
                only_longest_separate=only_longest_separate,
                polyfit_rst=polyfit_rst)

            map_figure, map_axis = plt.subplots(figsize=(10, 10))
            rst_map = plotRSTs_instance.create_map(map_axis, show_rst_info=True, req_colormap='coolwarm', show_info=show_info)

            if save_maps:
                # directory = 'C:/Users/hatzv/Documents/Geography/RSTs/python/Results/Test_success_rates/All_RSTs_1998-2000/'
                directory = 'C:/Users/hatzv/Documents/Geography/RSTs/python/Results/Test_Kuri/'
                map_name = current_day_str[0:10]
                filename = directory + map_name + ".png"
                plt.savefig(filename)
            else:
                plt.show()


            # if daily_rst_classification == consts.rst_orientation_no_rst:
            #     map_figure, map_axis = plt.subplots(figsize=(10, 10))
            #     rst_map = plotRSTs_instance.create_map(map_axis, show_rst_info=True, req_colormap='coolwarm')
            #
            #     if save_maps:
            #         directory = 'C:/Users/hatzv/Documents/Geography/RSTs/python/Results/Test_success_rates/No_RST_1998-2000/'
            #         map_name = current_day_str[0:10]
            #         filename = directory + map_name + ".png"
            #         plt.savefig(filename)
            #     else:
            #         plt.show()


