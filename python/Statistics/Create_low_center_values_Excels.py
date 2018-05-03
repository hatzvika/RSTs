from python.Plot_RSTs.Plot_RSTs import PlotRSTs
import numpy as np
from openpyxl import Workbook
from python.utils.find_low_center_in_area import find_low_center_in_area
import python.Plot_RSTs.plot_RST_constants as consts


NCEP_start_year = 1979  # Can be 1979 as well
NCEP_end_year = 2016

# Choose if only NCEP will be done
only_NCEP = False

# Parameters for the calculations that can change
use_interpolation = True
only_longest_separate = True
polyfit_rst = True

# Parameters for the calculations that can't change
data_to_map_var = 'Geostrophic Vorticity'
show_dots = False

excel_filename_NCEP = 'C:/Users/hatzv/Documents/Geography/RSTs/python/Statistics/Results/lowest_depth_one_dir_when_RST_found_NCEP_' + str(NCEP_start_year) + '-' + str(NCEP_end_year) + '_32-38_low_centers.xlsx'
excel_filename_ERA = 'C:/Users/hatzv/Documents/Geography/RSTs/python/Statistics/Results/lowest_depth_one_dir_when_RST_found_ERA_1979-2016_32-38_low_centers.xlsx'
excel_filename_ERA_2_5 = 'C:/Users/hatzv/Documents/Geography/RSTs/python/Statistics/Results/lowest_depth_one_dir_when_RST_found_ERA_2.5_1979-2016_32-38_low_centers.xlsx'

wb_NCEP = Workbook()
ws_NCEP = wb_NCEP.active
wb_ERA = Workbook()
ws_ERA = wb_ERA.active
wb_ERA_25 = Workbook()
ws_ERA_25 = wb_ERA_25.active

if only_NCEP:
    start_year = NCEP_start_year
    end_year = NCEP_end_year
else:
    start_year = 1979
    end_year = 2016

year_list = [str(x) for x in range(start_year, end_year+1)]

cols_counter = 2
for current_year in year_list:
    print(current_year)
    plotRSTs_NCEP_instance = PlotRSTs('NCEP', current_year)
    if not only_NCEP:
        plotRSTs_ERA_instance = PlotRSTs('ERA_Interim', current_year)
        plotRSTs_ERA_25_instance = PlotRSTs('ERA Int 2.5', current_year)
    data_string_time = plotRSTs_ERA_25_instance.data_string_time
    previous_month_day = ""
    leap_year_offset = 0
    rows_counter = 2
    for current_day in data_string_time:
        current_day_str = str(current_day)
        print(current_day_str)
        NCEP_daily_rst_classification,_,_,_ = plotRSTs_NCEP_instance.calculate_maps_data(current_day_str,
                                                                                         use_interpolation=use_interpolation,
                                                                                         data_to_map=data_to_map_var,
                                                                                         show_dots=show_dots,
                                                                                         only_longest_separate=only_longest_separate,
                                                                                         polyfit_rst=polyfit_rst)
        _, _, NCEP_low_center_depth1, NCEP_lowest_value_in_one_direction1 = find_low_center_in_area(plotRSTs_NCEP_instance.get_daily_slp_data(current_day_str),
                                                                                               plotRSTs_NCEP_instance.get_lats(),
                                                                                               plotRSTs_NCEP_instance.get_lons(),
                                                                                               consts.interp_resolution, 30, 35, 35, 42.5, 110, 300)
        _, _, NCEP_low_center_depth2, NCEP_lowest_value_in_one_direction2 = find_low_center_in_area(plotRSTs_NCEP_instance.get_daily_slp_data(current_day_str),
                                                                                               plotRSTs_NCEP_instance.get_lats(),
                                                                                               plotRSTs_NCEP_instance.get_lons(),
                                                                                               consts.interp_resolution, 25, 32.5, 25, 35, 110, 300)
        if not only_NCEP:
            ERA_daily_rst_classification,_,_,_ = plotRSTs_ERA_instance.calculate_maps_data(current_day_str,
                                                                                           use_interpolation=use_interpolation,
                                                                                           data_to_map=data_to_map_var,
                                                                                           show_dots=show_dots,
                                                                                           only_longest_separate=only_longest_separate,
                                                                                           polyfit_rst=polyfit_rst)
            _, _, ERA_low_center_depth1, ERA_lowest_value_in_one_direction1 = find_low_center_in_area(plotRSTs_ERA_instance.get_daily_slp_data(current_day_str),
                                                                                                  plotRSTs_ERA_instance.get_lats(),
                                                                                                  plotRSTs_ERA_instance.get_lons(),
                                                                                                  consts.interp_resolution, 30, 35, 35, 42.5, 110, 300)
            _, _, ERA_low_center_depth2, ERA_lowest_value_in_one_direction2 = find_low_center_in_area(plotRSTs_ERA_instance.get_daily_slp_data(current_day_str),
                                                                                                  plotRSTs_ERA_instance.get_lats(),
                                                                                                  plotRSTs_ERA_instance.get_lons(),
                                                                                                  consts.interp_resolution, 25, 32.5, 25, 35, 110, 300)
            ERA_25_daily_rst_classification,_,_,_ = plotRSTs_ERA_25_instance.calculate_maps_data(current_day_str,
                                                                                           use_interpolation=use_interpolation,
                                                                                           data_to_map=data_to_map_var,
                                                                                           show_dots=show_dots,
                                                                                           only_longest_separate=only_longest_separate,
                                                                                           polyfit_rst=polyfit_rst)
            _, _, ERA_25_low_center_depth1, ERA_25_lowest_value_in_one_direction1 = find_low_center_in_area(plotRSTs_ERA_25_instance.get_daily_slp_data(current_day_str),
                                                                    plotRSTs_ERA_25_instance.get_lats(),
                                                                    plotRSTs_ERA_25_instance.get_lons(),
                                                                    consts.interp_resolution, 30, 35, 35, 42.5, 110, 300)
            _, _, ERA_25_low_center_depth2, ERA_25_lowest_value_in_one_direction2 = find_low_center_in_area(plotRSTs_ERA_25_instance.get_daily_slp_data(current_day_str),
                                                                    plotRSTs_ERA_25_instance.get_lats(),
                                                                    plotRSTs_ERA_25_instance.get_lons(),
                                                                    consts.interp_resolution, 25, 32.5, 25, 35, 110, 300)

        current_month_day = current_day_str[5:10]
        if (previous_month_day == "02-28") and (current_month_day != "02-29"):
            leap_year_offset = 1

        # if (NCEP_daily_rst_classification == consts.rst_orientation_testing) and ((NCEP_low_center_depth1 > 0) or (NCEP_low_center_depth2 > 0)):
        #     ws_NCEP.cell(column=cols_counter, row=rows_counter+leap_year_offset, value=np.maximum(NCEP_low_center_depth1, NCEP_low_center_depth2))
        # if not only_NCEP:
        #     if (ERA_daily_rst_classification == consts.rst_orientation_testing) and ((ERA_low_center_depth1 > 0) or (ERA_low_center_depth2 > 0)):
        #         ws_ERA.cell(column=cols_counter, row=rows_counter + leap_year_offset, value=np.maximum(ERA_low_center_depth1, ERA_low_center_depth2))
        #     if (ERA_25_daily_rst_classification == consts.rst_orientation_testing) and ((ERA_25_low_center_depth1 > 0) or (ERA_25_low_center_depth2 > 0)):
        #         ws_ERA_25.cell(column=cols_counter, row=rows_counter + leap_year_offset, value=np.maximum(ERA_25_low_center_depth1, ERA_25_low_center_depth2))

        if (NCEP_daily_rst_classification != consts.rst_orientation_no_rst) and ((NCEP_low_center_depth1 > 0) or (NCEP_low_center_depth2 > 0)):
            if NCEP_low_center_depth1 > NCEP_low_center_depth2:
                ws_NCEP.cell(column=cols_counter, row=rows_counter+leap_year_offset, value=NCEP_lowest_value_in_one_direction1)
            else:
                ws_NCEP.cell(column=cols_counter, row=rows_counter + leap_year_offset, value=NCEP_lowest_value_in_one_direction2)
        if not only_NCEP:
            if (ERA_daily_rst_classification != consts.rst_orientation_no_rst) and ((ERA_low_center_depth1 > 0) or (ERA_low_center_depth2 > 0)):
                if ERA_low_center_depth1 > ERA_low_center_depth2:
                    ws_ERA.cell(column=cols_counter, row=rows_counter + leap_year_offset, value=ERA_lowest_value_in_one_direction1)
                else:
                    ws_ERA.cell(column=cols_counter, row=rows_counter + leap_year_offset, value=ERA_lowest_value_in_one_direction2)
            if (ERA_25_daily_rst_classification != consts.rst_orientation_no_rst) and ((ERA_25_low_center_depth1 > 0) or (ERA_25_low_center_depth2 > 0)):
                if ERA_25_low_center_depth1 > ERA_25_low_center_depth2:
                    ws_ERA_25.cell(column=cols_counter, row=rows_counter + leap_year_offset, value=ERA_25_lowest_value_in_one_direction1)
                else:
                    ws_ERA_25.cell(column=cols_counter, row=rows_counter + leap_year_offset, value=ERA_25_lowest_value_in_one_direction2)

        previous_month_day = current_month_day
        rows_counter = rows_counter + 1

    cols_counter = cols_counter + 1

wb_NCEP.save(excel_filename_NCEP)
if not only_NCEP:
    wb_ERA.save(excel_filename_ERA)
    wb_ERA_25.save(excel_filename_ERA_2_5)
