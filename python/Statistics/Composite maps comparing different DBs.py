import numpy as np
from openpyxl import load_workbook
from time import strptime
import matplotlib.pyplot as plt
from mpl_toolkits.basemap import Basemap

import python.Plot_RSTs.plot_RST_constants as consts
from python.Plot_RSTs.Plot_RSTs import PlotRSTs
from python.Plot_RSTs.Calculate_RST import Calculate_RST

#############################################
# First DB to create a composite map for
#############################################
first_req_DB = "NCEP"
# first_req_DB = "ERA"
# first_req_DB = "ERA_2_5"

#############################################
# Which classes to create a composite map for
#############################################
first_req_DB_class = "No RST"
# first_req_DB_class = "East"
# first_req_DB_class = "West"
# first_req_DB_class = "Central"

#############################################
# Second DB to create a composite map for
#############################################
# second_req_DB = "NCEP"
# second_req_DB = "ERA"
second_req_DB = "ERA_2_5"

#############################################
# Which classes to create a composite map for
#############################################
second_req_DB_class = "No RST"
# second_req_DB_class = "East"
# second_req_DB_class = "West"
# second_req_DB_class = "Central"
# second_req_DB_class = "ALL RSTs"

current_year = 1979 # This is the starting year. Later it will changed in the main loop

# Prepare the requested worksheets for the input classifications and initialize the plotRSTs objects according to the start year
if first_req_DB == "NCEP":
    first_excel_filename = 'C:/Users/hatzv/Documents/Geography/RSTs/python/Statistics/Results/RST_classification_NCEP_1979-2016.xlsx'
    first_plotRSTs_instance = PlotRSTs('NCEP', current_year)
elif first_req_DB == "ERA":
    first_excel_filename = 'C:/Users/hatzv/Documents/Geography/RSTs/python/Statistics/Results/RST_classification_ERA_1979-2016.xlsx'
    first_plotRSTs_instance = PlotRSTs('ERA_Interim', current_year)
else:  # ERA_2_5
    first_excel_filename = 'C:/Users/hatzv/Documents/Geography/RSTs/python/Statistics/Results/RST_classification_ERA_2.5_1979-2016.xlsx'
    first_plotRSTs_instance = PlotRSTs('ERA Int 2.5', current_year)

if second_req_DB == "NCEP":
    second_excel_filename = 'C:/Users/hatzv/Documents/Geography/RSTs/python/Statistics/Results/RST_classification_NCEP_1979-2016.xlsx'
    second_plotRSTs_instance = PlotRSTs('NCEP', current_year)
elif second_req_DB == "ERA":
    second_excel_filename = 'C:/Users/hatzv/Documents/Geography/RSTs/python/Statistics/Results/RST_classification_ERA_1979-2016.xlsx'
    second_plotRSTs_instance = PlotRSTs('ERA_Interim', current_year)
else:  # ERA_2_5
    second_excel_filename = 'C:/Users/hatzv/Documents/Geography/RSTs/python/Statistics/Results/RST_classification_ERA_2.5_1979-2016.xlsx'
    second_plotRSTs_instance = PlotRSTs('ERA Int 2.5', current_year)

first_wb = load_workbook(first_excel_filename, read_only=True)
first_ws = first_wb.active
first_table = first_ws['A1':'AM367']

second_wb = load_workbook(second_excel_filename, read_only=True)
second_ws = second_wb.active
second_table = second_ws['A1':'AM367']

# Isabella's classification is used only for getting the current day.
excel_filename_Isabella = 'C:/Users/hatzv/Documents/Geography/Research_help/synoptic_classification/My_classification/synoptic_classification_1948-21_Sep_2017.xlsx'
wb_Isabella = load_workbook(excel_filename_Isabella, read_only=True)
ws_Isabella = wb_Isabella.active
table_Isabella = ws_Isabella['A1':'BS367']

match_cases_counter = 0
first_is_rst_cond_met_counter = 0
second_is_rst_cond_met_counter = 0
first_composite_slp_map = np.zeros([81, 61])
first_composite_geostrophic_map = np.zeros([81, 61])
second_composite_slp_map = np.zeros([81, 61])
second_composite_geostrophic_map = np.zeros([81, 61])

# Main loop to collect all the single match map for the final composite map
for col in range (0, first_ws.max_column - 1):
    for row in range(0, first_ws.max_row - 1):
        first_value = first_table[row + 1][col + 1].value
        second_value = second_table[row + 1][col + 1].value

        if (first_value == first_req_DB_class): # and (second_value == second_req_DB_class):
            match_cases_counter = match_cases_counter + 1

            # Find the string date to pass to the PlotRST class
            current_year_str = str(first_table[0][col+1].value)
            current_month = strptime(first_table[row+1][0].value, '%b').tm_mon
            if current_month < 10:
                current_month_str = "0" + str(current_month)
            else:
                current_month_str = str(current_month)
            current_day = table_Isabella[row+1][1].value
            if current_day < 10:
                current_day_str = "0" + str(current_day)
            else:
                current_day_str = str(current_day)
            current_date = current_year_str + "-" + current_month_str + "-" + current_day_str + " 12:00:00"

            # Change the PlotRST instance for the current year, if needed
            if current_year < int(current_year_str):
                current_year = int(current_year_str)
                if first_req_DB == "NCEP":
                    first_plotRSTs_instance = PlotRSTs('NCEP', current_year)
                elif first_req_DB == "ERA":
                    first_plotRSTs_instance = PlotRSTs('ERA_Interim', current_year)
                else:  # ERA_2_5
                    first_plotRSTs_instance = PlotRSTs('ERA Int 2.5', current_year)

                if second_req_DB == "NCEP":
                    second_plotRSTs_instance = PlotRSTs('NCEP', current_year)
                elif second_req_DB == "ERA":
                    second_plotRSTs_instance = PlotRSTs('ERA_Interim', current_year)
                else:  # ERA_2_5
                    second_plotRSTs_instance = PlotRSTs('ERA Int 2.5', current_year)

            # Calculate the maps for the current date
            _rst_class, first_slp_map, first_geostrophic_vorticity_map, first_is_rst_conditions_met = first_plotRSTs_instance.calculate_maps_data(current_date,
                                                                                                   use_interpolation=True,
                                                                                                   data_to_map="Geostrophic Vorticity",
                                                                                                   show_dots=False,
                                                                                                   only_longest_separate=True,
                                                                                                   polyfit_rst=False)

            _rst_class, second_slp_map, second_geostrophic_vorticity_map, second_is_rst_conditions_met = second_plotRSTs_instance.calculate_maps_data(current_date,
                                                                                                   use_interpolation=True,
                                                                                                   data_to_map="Geostrophic Vorticity",
                                                                                                   show_dots=False,
                                                                                                   only_longest_separate=True,
                                                                                                   polyfit_rst=False)

            # Add the current date maps to the composite map
            first_composite_slp_map = first_composite_slp_map + first_slp_map
            first_composite_geostrophic_map = first_composite_geostrophic_map + first_geostrophic_vorticity_map
            second_composite_slp_map = second_composite_slp_map + second_slp_map
            second_composite_geostrophic_map = second_composite_geostrophic_map + second_geostrophic_vorticity_map

            if first_is_rst_conditions_met:
                first_is_rst_cond_met_counter = first_is_rst_cond_met_counter + 1
                print(current_date)
            if second_is_rst_conditions_met:
                second_is_rst_cond_met_counter = second_is_rst_cond_met_counter + 1

            # print(current_date)

print(match_cases_counter, first_is_rst_cond_met_counter, second_is_rst_cond_met_counter)

# Divide the composite by the number of cases to get the average values
first_composite_slp_map = first_composite_slp_map / match_cases_counter
first_composite_geostrophic_map = first_composite_geostrophic_map / match_cases_counter
second_composite_slp_map = second_composite_slp_map / match_cases_counter
second_composite_geostrophic_map = second_composite_geostrophic_map / match_cases_counter

# Display the maps
fig = plt.figure()

ax = fig.add_subplot(121)
# Calculate the meshes for the maps and plot SLP contours (always)
lons = first_plotRSTs_instance.interp_data_lons
lats = first_plotRSTs_instance.interp_data_lats

lon1_index = int(np.where(lons <= consts.map_lon1)[0][-1])
lon2_index = int(np.where(lons >= consts.map_lon2)[0][-1])
lat1_index = int(np.where(lats <= consts.map_lat1)[0][-1])
lat2_index = int(np.where(lats >= consts.map_lat2)[0][-1])
mesh_lons, mesh_lats = np.meshgrid(lons[lon1_index:lon2_index + 1], lats[lat1_index:lat2_index + 1])

# Create the first map object
first_rst_map = Basemap(llcrnrlon=consts.map_lon1,
                        llcrnrlat=consts.map_lat1,
                        urcrnrlon=consts.map_lon2,
                        urcrnrlat=consts.map_lat2,
                        projection='merc',
                        resolution='i')
first_rst_map.drawcoastlines()
first_rst_map.drawparallels(np.arange(consts.map_lat1, consts.map_lat2, 2.5), labels=[1, 0, 0, 0], fontsize=8)
first_rst_map.drawmeridians(np.arange(consts.map_lon1, consts.map_lon2, 2.5), labels=[0, 0, 0, 1], fontsize=8)
# map_axis.set(cmap = plt.cm.get_cmap('Blues_r')) #colormap = 'coolwarm',

# Build map title
plot_title = first_req_DB + ", " + first_req_DB_class + " for " + second_req_DB + ", " + second_req_DB_class
ax.set_title(plot_title)

x, y = first_rst_map(mesh_lons, mesh_lats)
min_slp = int(np.floor(first_composite_slp_map.min()))
max_slp = int(np.ceil(first_composite_slp_map.max()))
first_rst_map.contour(x, y, first_composite_slp_map[lat1_index:lat2_index + 1, lon1_index:lon2_index + 1], range(min_slp, max_slp, 200), linewidths=1.5,
                      colors='k')

# Display the geostrophic vorticity map if needed
subset_map = first_composite_geostrophic_map[lat1_index:lat2_index + 1, lon1_index:lon2_index + 1]

first_rst_map.contour(x, y, subset_map, 10, linewidths=0.5, colors='k')
max_value = subset_map.max()
min_value = subset_map.min()
if min_value < 0 and max_value > 0:
    if max_value > abs(min_value):
        min_value = -max_value
    else:
        max_value = abs(min_value)
cs = first_rst_map.contourf(x, y, subset_map, 10, cmap=plt.cm.get_cmap('coolwarm'), vmin=min_value, vmax=max_value)

#Draw the average RST
# Find Red Sea Trough
calc_rst_obj = Calculate_RST(first_composite_slp_map, 0.5, consts.slp_check_distance, lats, lons) # 0.5 is the interpolated resolution

# Find the RSTs orientations and polyfit them + get the daily_rst_classification (unless rst conditions are not met
trough_coordinates_matrix, stam, stam1 = calc_rst_obj.get_trough_coords_matrix(only_long_separate_RSTs=True, polyfit_rst=True)

# Draw the RSTs, if any
for current_RST in range(0, int(np.size(trough_coordinates_matrix, 1) / 2)):
    # Get the current trough columns from the trough matrix
    trough_coords = trough_coordinates_matrix[:, 2 * current_RST:2 * current_RST + 2]
    # remove all zeros from current RST
    trough_coords = trough_coords[~(trough_coords == 0).all(1)]
    x_trough = trough_coords[:, 1]
    y_trough = trough_coords[:, 0]

    lat_map, lon_map = first_rst_map(x_trough, y_trough)
    first_rst_map.plot(lat_map, lon_map, marker=None, linewidth=6, color='black')
    first_rst_map.plot(lat_map, lon_map, marker=None, linewidth=4, color='yellow')

# Draw box3 and the 2 points
lat_array_region = [consts.rst_square3_lat1,
                    consts.rst_square3_lat1,
                    consts.rst_square3_lat2,
                    consts.rst_square3_lat2,
                    consts.rst_square3_lat1]
lon_array_region = [consts.rst_square3_lon1,
                    consts.rst_square3_lon2,
                    consts.rst_square3_lon2,
                    consts.rst_square3_lon1,
                    consts.rst_square3_lon1]
x_region, y_region = first_rst_map(lon_array_region, lat_array_region)
first_rst_map.plot(x_region, y_region, marker=None, linewidth=3, color='black')

# Draw the central red line
x_line, y_line = first_rst_map([consts.central_cross_line_lon, consts.central_cross_line_lon],
                               [consts.central_cross_line_lat1, consts.central_cross_line_lat2])
first_rst_map.plot(x_line, y_line, marker=None, linewidth=6, color='black')
first_rst_map.plot(x_line, y_line, marker=None, linewidth=3, color='red')

# Add Colorbar
plt.colorbar(cs)

# Create the second map object
ax = fig.add_subplot(122)
# Calculate the meshes for the maps and plot SLP contours (always)
lons = first_plotRSTs_instance.interp_data_lons
lats = first_plotRSTs_instance.interp_data_lats

lon1_index = int(np.where(lons <= consts.map_lon1)[0][-1])
lon2_index = int(np.where(lons >= consts.map_lon2)[0][-1])
lat1_index = int(np.where(lats <= consts.map_lat1)[0][-1])
lat2_index = int(np.where(lats >= consts.map_lat2)[0][-1])
mesh_lons, mesh_lats = np.meshgrid(lons[lon1_index:lon2_index + 1], lats[lat1_index:lat2_index + 1])

second_rst_map = Basemap(llcrnrlon=consts.map_lon1,
                        llcrnrlat=consts.map_lat1,
                        urcrnrlon=consts.map_lon2,
                        urcrnrlat=consts.map_lat2,
                        projection='merc',
                        resolution='i')
second_rst_map.drawcoastlines()
second_rst_map.drawparallels(np.arange(consts.map_lat1, consts.map_lat2, 2.5), labels=[1, 0, 0, 0], fontsize=8)
second_rst_map.drawmeridians(np.arange(consts.map_lon1, consts.map_lon2, 2.5), labels=[0, 0, 0, 1], fontsize=8)
# map_axis.set(cmap = plt.cm.get_cmap('Blues_r')) #colormap = 'coolwarm',

# Build map title
plot_title = second_req_DB + ", " + second_req_DB_class + " for " + first_req_DB + ", " + first_req_DB_class
ax.set_title(plot_title)

x, y = second_rst_map(mesh_lons, mesh_lats)
min_slp = int(np.floor(second_composite_slp_map.min()))
max_slp = int(np.ceil(second_composite_slp_map.max()))
second_rst_map.contour(x, y, second_composite_slp_map[lat1_index:lat2_index + 1, lon1_index:lon2_index + 1], range(min_slp, max_slp, 200), linewidths=1.5,
                      colors='k')

# Display the geostrophic vorticity map if needed
subset_map = second_composite_geostrophic_map[lat1_index:lat2_index + 1, lon1_index:lon2_index + 1]

second_rst_map.contour(x, y, subset_map, 10, linewidths=0.5, colors='k')
max_value = subset_map.max()
min_value = subset_map.min()
if min_value < 0 and max_value > 0:
    if max_value > abs(min_value):
        min_value = -max_value
    else:
        max_value = abs(min_value)
cs = second_rst_map.contourf(x, y, subset_map, 10, cmap=plt.cm.get_cmap('coolwarm'), vmin=min_value, vmax=max_value)

#Draw the average RST
# Find Red Sea Trough
calc_rst_obj = Calculate_RST(second_composite_slp_map, 0.5, consts.slp_check_distance, lats, lons) # 0.5 is the interpolated resolution

# Find the RSTs orientations and polyfit them + get the daily_rst_classification (unless rst conditions are not met
trough_coordinates_matrix, stam, stam1 = calc_rst_obj.get_trough_coords_matrix(only_long_separate_RSTs=True, polyfit_rst=True)

# Draw the RSTs, if any
for current_RST in range(0, int(np.size(trough_coordinates_matrix, 1) / 2)):
    # Get the current trough columns from the trough matrix
    trough_coords = trough_coordinates_matrix[:, 2 * current_RST:2 * current_RST + 2]
    # remove all zeros from current RST
    trough_coords = trough_coords[~(trough_coords == 0).all(1)]
    x_trough = trough_coords[:, 1]
    y_trough = trough_coords[:, 0]

    lat_map, lon_map = second_rst_map(x_trough, y_trough)
    second_rst_map.plot(lat_map, lon_map, marker=None, linewidth=6, color='black')
    second_rst_map.plot(lat_map, lon_map, marker=None, linewidth=4, color='yellow')

# Draw box3 and the 2 points
lat_array_region = [consts.rst_square3_lat1,
                    consts.rst_square3_lat1,
                    consts.rst_square3_lat2,
                    consts.rst_square3_lat2,
                    consts.rst_square3_lat1]
lon_array_region = [consts.rst_square3_lon1,
                    consts.rst_square3_lon2,
                    consts.rst_square3_lon2,
                    consts.rst_square3_lon1,
                    consts.rst_square3_lon1]
x_region, y_region = second_rst_map(lon_array_region, lat_array_region)
second_rst_map.plot(x_region, y_region, marker=None, linewidth=3, color='black')

# Draw the central red line
x_line, y_line = second_rst_map([consts.central_cross_line_lon, consts.central_cross_line_lon],
                               [consts.central_cross_line_lat1, consts.central_cross_line_lat2])
second_rst_map.plot(x_line, y_line, marker=None, linewidth=6, color='black')
second_rst_map.plot(x_line, y_line, marker=None, linewidth=3, color='red')

# Add Colorbar
plt.colorbar(cs)
plt.show()


