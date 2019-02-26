from python.Plot_RSTs.Plot_RSTs import PlotRSTs
#import numpy as np
from openpyxl import Workbook

NCEP_start_year = 1948  # Can be 1979 as well
NCEP_end_year = 2017

# Choose if only NCEP will be done
only_NCEP = False

# Parameters for the calculations that can change
use_interpolation = True
only_longest_separate = True
polyfit_rst = False

# Parameters for the calculations that can't change
data_to_map_var = 'Geostrophic Vorticity'
show_dots = False

excel_filename_NCEP = 'C:/Users/hatzv/Documents/Geography/RSTs/python/Analysis/Results/RST_classification_all_hours_NCEP_' + str(NCEP_start_year) + '-' + str(NCEP_end_year) + '.xlsx'
excel_filename_ERA = 'C:/Users/hatzv/Documents/Geography/RSTs/python/Analysis/Results/RST_classification_all_hours_ERA_1979-2016.xlsx'
excel_filename_ERA_2_5 = 'C:/Users/hatzv/Documents/Geography/RSTs/python/Analysis/Results/RST_classification_all_hours_ERA_2.5_1979-2016.xlsx'

wb_NCEP = Workbook()
ws_NCEP = wb_NCEP.active
wb_ERA = Workbook()
ws_ERA = wb_ERA.active
wb_ERA_25 = Workbook()
ws_ERA_25 = wb_ERA_25.active

if only_NCEP:
    start_year = NCEP_start_year
    end_year = NCEP_end_year
else:
    start_year = 1979
    end_year = 2016
    # end_year = 2000

year_list = [str(x) for x in range(start_year, end_year+1)]

cols_counter = 2
for current_year in year_list:
    print(current_year)
    plotRSTs_NCEP_instance_00 = PlotRSTs('NCEP', current_year, z_hour=0)
    plotRSTs_NCEP_instance_06 = PlotRSTs('NCEP', current_year, z_hour=6)
    plotRSTs_NCEP_instance_12 = PlotRSTs('NCEP', current_year, z_hour=12)
    plotRSTs_NCEP_instance_18 = PlotRSTs('NCEP', current_year, z_hour=18)
    if not only_NCEP:
        plotRSTs_ERA_instance_00 = PlotRSTs('ERA_Interim', current_year, z_hour=0)
        plotRSTs_ERA_instance_06 = PlotRSTs('ERA_Interim', current_year, z_hour=6)
        plotRSTs_ERA_instance_12 = PlotRSTs('ERA_Interim', current_year, z_hour=12)
        plotRSTs_ERA_instance_18 = PlotRSTs('ERA_Interim', current_year, z_hour=18)
        plotRSTs_ERA_25_instance_00 = PlotRSTs('ERA Int 2.5', current_year, z_hour=0)
        plotRSTs_ERA_25_instance_06 = PlotRSTs('ERA Int 2.5', current_year, z_hour=6)
        plotRSTs_ERA_25_instance_12 = PlotRSTs('ERA Int 2.5', current_year, z_hour=12)
        plotRSTs_ERA_25_instance_18 = PlotRSTs('ERA Int 2.5', current_year, z_hour=18)
    data_string_time = plotRSTs_NCEP_instance_00.data_string_time
    previous_month_day = ""
    leap_year_offset = 0
    rows_counter = 2
    for current_day in data_string_time:
        current_day_list = list(str(current_day))
        current_day_str = ''.join(current_day_list)
        print(current_day_str)
        # Find the RSTs for 00z
        NCEP_daily_rst_classification_00,_,_,_ = plotRSTs_NCEP_instance_00.calculate_maps_data(current_day_str,
                                                                                   use_interpolation=use_interpolation,
                                                                                   data_to_map=data_to_map_var,
                                                                                   show_dots=show_dots,
                                                                                   only_longest_separate=only_longest_separate,
                                                                                   polyfit_rst=polyfit_rst)
        if not only_NCEP:
            ERA_daily_rst_classification_00,_,_,_ = plotRSTs_ERA_instance_00.calculate_maps_data(current_day_str,
                                                                                     use_interpolation=use_interpolation,
                                                                                     data_to_map=data_to_map_var,
                                                                                     show_dots=show_dots,
                                                                                     only_longest_separate=only_longest_separate,
                                                                                     polyfit_rst=polyfit_rst)
            ERA_25_daily_rst_classification_00,_,_,_ = plotRSTs_ERA_25_instance_00.calculate_maps_data(current_day_str,
                                                                                           use_interpolation=use_interpolation,
                                                                                           data_to_map=data_to_map_var,
                                                                                           show_dots=show_dots,
                                                                                           only_longest_separate=only_longest_separate,
                                                                                           polyfit_rst=polyfit_rst)

        current_month_day = current_day_str[5:10]
        if (previous_month_day == "02-28") and (current_month_day != "02-29"):
            leap_year_offset = 4

        ws_NCEP.cell(column=cols_counter, row=rows_counter+leap_year_offset, value=NCEP_daily_rst_classification_00)
        if not only_NCEP:
            ws_ERA.cell(column=cols_counter, row=rows_counter+leap_year_offset, value=ERA_daily_rst_classification_00)
            ws_ERA_25.cell(column=cols_counter, row=rows_counter+leap_year_offset, value=ERA_25_daily_rst_classification_00)
        rows_counter = rows_counter + 1

        # Find the RSTs for 06z
        current_day_list[11:13] = '06'
        current_day_str = ''.join(current_day_list)
        NCEP_daily_rst_classification_06,_,_,_ = plotRSTs_NCEP_instance_06.calculate_maps_data(current_day_str,
                                                                                   use_interpolation=use_interpolation,
                                                                                   data_to_map=data_to_map_var,
                                                                                   show_dots=show_dots,
                                                                                   only_longest_separate=only_longest_separate,
                                                                                   polyfit_rst=polyfit_rst)
        if not only_NCEP:
            ERA_daily_rst_classification_06,_,_,_ = plotRSTs_ERA_instance_06.calculate_maps_data(current_day_str,
                                                                                     use_interpolation=use_interpolation,
                                                                                     data_to_map=data_to_map_var,
                                                                                     show_dots=show_dots,
                                                                                     only_longest_separate=only_longest_separate,
                                                                                     polyfit_rst=polyfit_rst)
            ERA_25_daily_rst_classification_06,_,_,_ = plotRSTs_ERA_25_instance_06.calculate_maps_data(current_day_str,
                                                                                           use_interpolation=use_interpolation,
                                                                                           data_to_map=data_to_map_var,
                                                                                           show_dots=show_dots,
                                                                                           only_longest_separate=only_longest_separate,
                                                                                           polyfit_rst=polyfit_rst)
        ws_NCEP.cell(column=cols_counter, row=rows_counter+leap_year_offset, value=NCEP_daily_rst_classification_06)
        if not only_NCEP:
            ws_ERA.cell(column=cols_counter, row=rows_counter+leap_year_offset, value=ERA_daily_rst_classification_06)
            ws_ERA_25.cell(column=cols_counter, row=rows_counter+leap_year_offset, value=ERA_25_daily_rst_classification_06)
        rows_counter = rows_counter + 1

        # Find the RSTs for 12z
        current_day_list[11:13] = '12'
        current_day_str = ''.join(current_day_list)
        NCEP_daily_rst_classification_12,_,_,_ = plotRSTs_NCEP_instance_12.calculate_maps_data(current_day_str,
                                                                                   use_interpolation=use_interpolation,
                                                                                   data_to_map=data_to_map_var,
                                                                                   show_dots=show_dots,
                                                                                   only_longest_separate=only_longest_separate,
                                                                                   polyfit_rst=polyfit_rst)
        if not only_NCEP:
            ERA_daily_rst_classification_12,_,_,_ = plotRSTs_ERA_instance_12.calculate_maps_data(current_day_str,
                                                                                     use_interpolation=use_interpolation,
                                                                                     data_to_map=data_to_map_var,
                                                                                     show_dots=show_dots,
                                                                                     only_longest_separate=only_longest_separate,
                                                                                     polyfit_rst=polyfit_rst)
            ERA_25_daily_rst_classification_12,_,_,_ = plotRSTs_ERA_25_instance_12.calculate_maps_data(current_day_str,
                                                                                           use_interpolation=use_interpolation,
                                                                                           data_to_map=data_to_map_var,
                                                                                           show_dots=show_dots,
                                                                                           only_longest_separate=only_longest_separate,
                                                                                           polyfit_rst=polyfit_rst)
        ws_NCEP.cell(column=cols_counter, row=rows_counter+leap_year_offset, value=NCEP_daily_rst_classification_12)
        if not only_NCEP:
            ws_ERA.cell(column=cols_counter, row=rows_counter+leap_year_offset, value=ERA_daily_rst_classification_12)
            ws_ERA_25.cell(column=cols_counter, row=rows_counter+leap_year_offset, value=ERA_25_daily_rst_classification_12)
        rows_counter = rows_counter + 1

        # Find the RSTs for 18z
        current_day_list[11:13] = '18'
        current_day_str = ''.join(current_day_list)
        NCEP_daily_rst_classification_18,_,_,_ = plotRSTs_NCEP_instance_18.calculate_maps_data(current_day_str,
                                                                                   use_interpolation=use_interpolation,
                                                                                   data_to_map=data_to_map_var,
                                                                                   show_dots=show_dots,
                                                                                   only_longest_separate=only_longest_separate,
                                                                                   polyfit_rst=polyfit_rst)
        if not only_NCEP:
            ERA_daily_rst_classification_18,_,_,_ = plotRSTs_ERA_instance_18.calculate_maps_data(current_day_str,
                                                                                     use_interpolation=use_interpolation,
                                                                                     data_to_map=data_to_map_var,
                                                                                     show_dots=show_dots,
                                                                                     only_longest_separate=only_longest_separate,
                                                                                     polyfit_rst=polyfit_rst)
            ERA_25_daily_rst_classification_18,_,_,_ = plotRSTs_ERA_25_instance_18.calculate_maps_data(current_day_str,
                                                                                           use_interpolation=use_interpolation,
                                                                                           data_to_map=data_to_map_var,
                                                                                           show_dots=show_dots,
                                                                                           only_longest_separate=only_longest_separate,
                                                                                           polyfit_rst=polyfit_rst)
        ws_NCEP.cell(column=cols_counter, row=rows_counter+leap_year_offset, value=NCEP_daily_rst_classification_18)
        if not only_NCEP:
            ws_ERA.cell(column=cols_counter, row=rows_counter+leap_year_offset, value=ERA_daily_rst_classification_18)
            ws_ERA_25.cell(column=cols_counter, row=rows_counter+leap_year_offset, value=ERA_25_daily_rst_classification_18)

        previous_month_day = current_month_day
        rows_counter = rows_counter + 1

    cols_counter = cols_counter + 1

wb_NCEP.save(excel_filename_NCEP)
if not only_NCEP:
    wb_ERA.save(excel_filename_ERA)
    wb_ERA_25.save(excel_filename_ERA_2_5)
