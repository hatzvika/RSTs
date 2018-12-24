from openpyxl import Workbook, load_workbook
from time import strptime

seasonal_trends = False  # False for monthly trends

output_dir = 'C:/Users/hatzv/Documents/Geography/RSTs/python/Analysis/Results/'

# Prepare the worksheets for the NCEP output file
if seasonal_trends:
    output_excel_filename_NCEP = output_dir + 'Seasonal_trends_RSTs_NCEP_1948-2017.xlsx'
else:
    output_excel_filename_NCEP = output_dir + 'Monthly_trends_RSTs_NCEP_1948-2017.xlsx'
wb_NCEP_trends = Workbook()
ws_NCEP_NO_RSTs = wb_NCEP_trends.get_sheet_by_name("Sheet")
ws_NCEP_NO_RSTs.title = "NO_RSTs"
ws_NCEP_East = wb_NCEP_trends.create_sheet("East")
ws_NCEP_West = wb_NCEP_trends.create_sheet("West")
ws_NCEP_Central = wb_NCEP_trends.create_sheet("Central")
ws_NCEP_All = wb_NCEP_trends.create_sheet("All RSTs")

# Prepare the output tables for all created sheets
table_NCEP_NO_RSTs = ws_NCEP_NO_RSTs['B2':'BS13']
table_NCEP_East = ws_NCEP_East['B2':'BS13']
table_NCEP_West = ws_NCEP_West['B2':'BS13']
table_NCEP_Central = ws_NCEP_Central['B2':'BS13']
table_NCEP_All = ws_NCEP_All['B2':'BS13']

# Prepare the worksheets for the input classifications
excel_filename_NCEP = 'C:/Users/hatzv/Documents/Geography/RSTs/python/Analysis/Results/RST_classification_NCEP_1948-2017.xlsx'

wb_NCEP = load_workbook(excel_filename_NCEP, read_only=True)
ws_NCEP = wb_NCEP.active
table_NCEP = ws_NCEP['A2':'BS367']

for row in range(1, ws_NCEP.max_row-1):
    for col in range(1, ws_NCEP.max_column):
        any_NCEP_RST = True  # A flag for marking if an RST found or not

        # Find the output row
        current_month = strptime(table_NCEP[row][0].value, '%b').tm_mon
        if seasonal_trends:
            if current_month == 12 or current_month <=2:  # DJF
                current_row = 0
            elif current_month >= 3 and current_month <=5:  # MAM
                current_row = 1
            elif current_month >= 6 and current_month <= 8:  # JJA
                current_row = 2
            else:  # SON
                current_row = 3
        else:
            current_row = current_month - 1

        NCEP_value = table_NCEP[row][col].value

        if NCEP_value == "No RST":
            output_NCEP_table = table_NCEP_NO_RSTs
            any_NCEP_RST = False
        elif NCEP_value == "East":
            output_NCEP_table = table_NCEP_East
        elif NCEP_value == "Central":
            output_NCEP_table = table_NCEP_Central
        elif NCEP_value == "West":
            output_NCEP_table = table_NCEP_West

        # TODO: Refactor this entire ugly thing.
        if NCEP_value is not None:  # Make sure not to read the None value in a non-leap year
            current_output_value_NCEP = output_NCEP_table[current_row][col-1].value
            if current_output_value_NCEP is not None:
                output_NCEP_table[current_row][col - 1].value = current_output_value_NCEP + 1
            else:
                output_NCEP_table[current_row][col - 1].value = 1

            if any_NCEP_RST:
                output_NCEP_table = table_NCEP_All
                current_output_value_NCEP = output_NCEP_table[current_row][col-1].value
                if current_output_value_NCEP is not None:
                    output_NCEP_table[current_row][col - 1].value = current_output_value_NCEP + 1
                else:
                    output_NCEP_table[current_row][col - 1].value = 1

wb_NCEP_trends.save(output_excel_filename_NCEP)
